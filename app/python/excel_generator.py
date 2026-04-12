# app/python/excel_generator.py
"""
Generates comprehensive Excel practice/example files for each course level.
Supports Spanish (es) and English (en) formula names.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Formula name maps ─────────────────────────────────────────────────────────

FORMULAS = {
    "es": {
        "SUM":        "SUMA",
        "IF":         "SI",
        "IFERROR":    "SI.ERROR",
        "IFS":        "SI.CONJUNTO",
        "AND":        "Y",
        "OR":         "O",
        "NOT":        "NO",
        "VLOOKUP":    "BUSCARV",
        "HLOOKUP":    "BUSCARH",
        "XLOOKUP":    "BUSCARX",
        "INDEX":      "INDICE",
        "MATCH":      "COINCIDIR",
        "COUNTIF":    "CONTAR.SI",
        "COUNTIFS":   "CONTAR.SI.CONJUNTO",
        "SUMIF":      "SUMAR.SI",
        "SUMIFS":     "SUMAR.SI.CONJUNTO",
        "AVERAGEIF":  "PROMEDIO.SI",
        "AVERAGEIFS": "PROMEDIO.SI.CONJUNTO",
        "COUNT":      "CONTAR",
        "COUNTA":     "CONTARA",
        "AVERAGE":    "PROMEDIO",
        "MAX":        "MAX",
        "MIN":        "MIN",
        "LEFT":       "IZQUIERDA",
        "RIGHT":      "DERECHA",
        "MID":        "EXTRAE",
        "LEN":        "LARGO",
        "TRIM":       "ESPACIOS",
        "UPPER":      "MAYUSC",
        "LOWER":      "MINUSC",
        "PROPER":     "NOMPROPIO",
        "CONCATENATE":"CONCATENAR",
        "TEXT":       "TEXTO",
        "VALUE":      "VALOR",
        "DATE":       "FECHA",
        "TODAY":      "HOY",
        "NOW":        "AHORA",
        "YEAR":       "AÑO",
        "MONTH":      "MES",
        "DAY":        "DIA",
        "DATEDIF":    "SIFECHA",
        "NETWORKDAYS":"DIAS.LAB",
        "EOMONTH":    "FIN.MES",
        "FILTER":     "FILTRAR",
        "SORT":       "ORDENAR",
        "UNIQUE":     "UNICOS",
        "INDIRECT":   "INDIRECTO",
        "OFFSET":     "DESREF",
        "ROWS":       "FILAS",
        "COLUMNS":    "COLUMNAS",
        "ROUND":      "REDONDEAR",
        "INT":        "ENTERO",
        "ABS":        "ABS",
        "SQRT":       "RAIZ",
        "POWER":      "POTENCIA",
        "MOD":        "RESIDUO",
        "FIND":       "ENCONTRAR",
        "SEARCH":     "HALLAR",
        "SUBSTITUTE": "SUSTITUIR",
        "REPLACE":    "REEMPLAZAR",
        "REPT":       "REPETIR",
        "EXACT":      "IGUAL",
        "TRANSPOSE":  "TRANSPONER",
        "LARGE":      "K.ESIMO.MAYOR",
        "SMALL":      "K.ESIMO.MENOR",
        "RANK":       "JERARQUIA",
        "ISBLANK":    "ESBLANCO",
        "ISNUMBER":   "ESNUMERO",
        "ISTEXT":     "ESTEXTO",
        "ISERROR":    "ESERROR",
    },
    "en": {
        k: k for k in [
            "SUM","IF","IFERROR","IFS","AND","OR","NOT","VLOOKUP","HLOOKUP",
            "XLOOKUP","INDEX","MATCH","COUNTIF","COUNTIFS","SUMIF","SUMIFS",
            "AVERAGEIF","AVERAGEIFS","COUNT","COUNTA","AVERAGE","MAX","MIN",
            "LEFT","RIGHT","MID","LEN","TRIM","UPPER","LOWER","PROPER",
            "CONCATENATE","TEXT","VALUE","DATE","TODAY","NOW","YEAR","MONTH",
            "DAY","DATEDIF","NETWORKDAYS","EOMONTH","FILTER","SORT","UNIQUE",
            "INDIRECT","OFFSET","ROWS","COLUMNS","ROUND","INT","ABS","SQRT",
            "POWER","MOD","FIND","SEARCH","SUBSTITUTE","REPLACE","REPT",
            "EXACT","TRANSPOSE","LARGE","SMALL","RANK","ISBLANK","ISNUMBER",
            "ISTEXT","ISERROR",
        ]
    }
}


def f(lang, name):
    """Return formula name in the requested language."""
    return FORMULAS[lang].get(name, name)


# ── Style helpers ─────────────────────────────────────────────────────────────

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def normal_font(size=11):
    return Font(size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def row_height(ws, row_num, height):
    ws.row_dimensions[row_num].height = height

def header_row(ws, row, values, fill_hex, font_color="FFFFFF", row_h=22):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = header_fill(fill_hex)
        cell.font = bold_font(11, font_color)
        cell.alignment = center()
        cell.border = thin_border()
    row_height(ws, row, row_h)

def data_cell(ws, row, col, value, bold=False, center_align=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = bold_font(11) if bold else normal_font(11)
    cell.alignment = center() if center_align else left()
    cell.border = thin_border()
    return cell

def section_title(ws, row, col, text, hex_color="1F3864", colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = header_fill(hex_color)
    cell.font = bold_font(13, "FFFFFF")
    cell.alignment = left()
    row_height(ws, row, 24)
    return cell

def note_cell(ws, row, col, text, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor="FFF9C4")
    cell.font = Font(italic=True, size=10, color="5D4037")
    cell.alignment = left()
    return cell

def formula_example_row(ws, row, description, formula_text, result_formula,
                         fill_hex="EEF2FF"):
    ws.cell(row=row, column=1, value=description).border = thin_border()
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill_hex)
    ws.cell(row=row, column=1).font = normal_font(10)
    ws.cell(row=row, column=2, value=formula_text).border = thin_border()
    ws.cell(row=row, column=2).font = Font(name="Courier New", size=10, color="1A237E")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="F8F9FA")
    c = ws.cell(row=row, column=3)
    if result_formula:
        c.value = result_formula
    c.border = thin_border()
    c.fill = PatternFill("solid", fgColor="F1F8E9")
    row_height(ws, row, 18)


# ══════════════════════════════════════════════════════════════════════════════
# NIVEL 0 — Fundamentos
# ══════════════════════════════════════════════════════════════════════════════

def _nivel0_ejemplos(wb, lang):
    # ── Hoja 1: Interfaz ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "01 Interfaz"
    ws.sheet_view.showGridLines = True

    section_title(ws, 1, 1, "📌 MÓDULO 1 — La Interfaz de Excel 365", "1F3864")
    note_cell(ws, 2, 1, "Este mapa te ayuda a ubicar los elementos principales de la pantalla de Excel.")

    headers = ["Zona de la interfaz", "¿Dónde está?", "Para qué sirve", "Atajo rápido"]
    header_row(ws, 4, headers, "1565C0")

    rows = [
        ("Barra de fórmulas", "Arriba del área de celdas", "Escribe o edita el contenido de una celda", "F2 para editar"),
        ("Cuadro de nombres", "Izquierda de la barra de fórmulas", "Muestra la dirección de celda activa o nombre de rango", "Ctrl+G para Ir a"),
        ("Cinta de opciones (Ribbon)", "Parte superior", "Acceso a todas las funciones organizadas por fichas", "Alt para ver atajos"),
        ("Barra de estado", "Parte inferior", "Muestra SUMA, PROMEDIO, CUENTA del rango seleccionado", "Clic derecho para personalizar"),
        ("Selector todo (all)", "Esquina sup. izquierda de celdas", "Selecciona TODAS las celdas de la hoja", "Ctrl+A"),
        ("Pestañas de hojas", "Parte inferior izquierda", "Navegar entre hojas del libro", "Ctrl+AvPág / Ctrl+RePág"),
        ("Vista Backstage", "Pestaña Archivo", "Guardar, abrir, exportar, opciones de Excel", "Alt+A o Alt+F"),
    ]
    for i, r in enumerate(rows, 5):
        for j, v in enumerate(r, 1):
            c = data_cell(ws, i, j, v)
        row_height(ws, i, 20)

    col_width(ws, "A", 30); col_width(ws, "B", 28); col_width(ws, "C", 38); col_width(ws, "D", 22)

    # ── Hoja 2: Atajos esenciales ─────────────────────────────────────────────
    ws2 = wb.create_sheet("02 Atajos")
    section_title(ws2, 1, 1, "⌨️ MÓDULO 2 — Atajos Esenciales de Excel 365", "1F3864")
    note_cell(ws2, 2, 1, "Practica estos atajos hasta que sean automáticos. 20 min/día durante 2 semanas = experto.")

    cats = [
        ("NAVEGACIÓN", "1565C0", [
            ("Ctrl + Inicio", "Ir a celda A1"),
            ("Ctrl + Fin", "Ir a la última celda con datos"),
            ("Ctrl + →  /  ←", "Saltar al borde del bloque de datos"),
            ("Ctrl + ↑  /  ↓", "Subir/bajar al borde del bloque"),
            ("Ctrl + AvPág", "Ir a la siguiente hoja"),
            ("Ctrl + RePág", "Ir a la hoja anterior"),
        ]),
        ("SELECCIÓN", "2E7D32", [
            ("Ctrl + Shift + Fin", "Seleccionar hasta última celda con datos"),
            ("Ctrl + Shift + Inicio", "Seleccionar hasta A1"),
            ("Ctrl + Shift + →", "Seleccionar hasta borde de bloque"),
            ("Ctrl + A", "Seleccionar toda la hoja"),
            ("Ctrl + Espacio", "Seleccionar columna completa"),
            ("Shift + Espacio", "Seleccionar fila completa"),
        ]),
        ("EDICIÓN", "E65100", [
            ("F2", "Entrar en modo edición de celda"),
            ("Ctrl + Z", "Deshacer"),
            ("Ctrl + Y", "Rehacer"),
            ("Ctrl + D", "Copiar celda de arriba hacia abajo"),
            ("Ctrl + R", "Copiar celda de la izquierda hacia la derecha"),
            ("Ctrl + ;", "Insertar fecha actual"),
            ("Ctrl + Shift + ;", "Insertar hora actual"),
            ("Alt + Enter", "Salto de línea dentro de la celda"),
            ("Delete", "Borrar contenido (mantiene formato)"),
        ]),
        ("FORMATO", "4A148C", [
            ("Ctrl + N", "Negrita"),
            ("Ctrl + K", "Cursiva"),
            ("Ctrl + S", "Subrayado"),
            ("Ctrl + 1", "Abrir cuadro Formato de celdas"),
            ("Ctrl + Shift + $", "Formato moneda"),
            ("Ctrl + Shift + %", "Formato porcentaje"),
            ("Ctrl + Shift + #", "Formato fecha"),
        ]),
        ("FÓRMULAS / LIBRO", "C62828", [
            ("= (igual)", "Iniciar una fórmula"),
            ("F4", "Cambiar tipo de referencia ($A$1 → A$1 → $A1 → A1)"),
            ("Ctrl + `", "Mostrar/ocultar fórmulas"),
            ("F9", "Calcular/actualizar fórmulas"),
            ("Alt + =", "Insertar SUMA automática"),
            ("Ctrl + T", "Crear tabla"),
            ("Ctrl + Shift + L", "Activar/desactivar filtros"),
        ]),
    ]

    r = 4
    for cat_name, cat_color, shortcuts in cats:
        header_row(ws2, r, [cat_name, "Acción"], cat_color)
        r += 1
        for atajo, desc in shortcuts:
            data_cell(ws2, r, 1, atajo, bold=True, center_align=True)
            data_cell(ws2, r, 2, desc)
            row_height(ws2, r, 18)
            r += 1
        r += 1

    col_width(ws2, "A", 28); col_width(ws2, "B", 50)

    # ── Hoja 3: Estructura de datos ───────────────────────────────────────────
    ws3 = wb.create_sheet("03 Estructura Datos")
    section_title(ws3, 1, 1, "🗂️ MÓDULO 3 — Estructura de Datos Correcta", "1F3864")
    note_cell(ws3, 2, 1, "Una tabla bien estructurada tiene: 1 fila de encabezados, 1 dato por celda, sin filas/columnas vacías.")

    header_row(ws3, 4, ["✅ CORRECTO — Registro de Ventas"], "2E7D32")
    header_row(ws3, 5, ["ID","Fecha","Vendedor","Producto","Categoría","Cantidad","Precio Unit.","Total"], "388E3C", "FFFFFF")

    sample = [
        (1,"2024-01-15","Ana García","Laptop Pro","Electrónicos",2,850.00),
        (2,"2024-01-16","Carlos López","Mouse Ergonómico","Periféricos",5,45.00),
        (3,"2024-01-17","María Torres","Teclado Mecánico","Periféricos",3,120.00),
        (4,"2024-01-18","Ana García","Monitor 27\"","Electrónicos",1,320.00),
        (5,"2024-01-19","Luis Morales","Webcam HD","Periféricos",4,75.00),
        (6,"2024-01-20","Carlos López","Silla Gamer","Mobiliario",2,280.00),
    ]
    for i, row_data in enumerate(sample, 6):
        for j, v in enumerate(row_data, 1):
            data_cell(ws3, i, j, v, center_align=(j in [1,3,4,5,6]))
        ws3.cell(row=i, column=8, value=f"=F{i}*G{i}").border = thin_border()
        row_height(ws3, i, 18)

    col_width(ws3, "A", 6); col_width(ws3, "B", 14); col_width(ws3, "C", 18)
    col_width(ws3, "D", 22); col_width(ws3, "E", 16); col_width(ws3, "F", 12)
    col_width(ws3, "G", 14); col_width(ws3, "H", 12)

    ws3.cell(row=13, column=1, value="❌ ERRORES COMUNES A EVITAR:").font = bold_font(11, "C62828")
    errors = [
        "• Celdas combinadas en encabezados (rompen filtros y fórmulas)",
        "• Filas totales dentro de la tabla (interfieren con tablas dinámicas)",
        "• Datos mixtos en una columna (números + texto = fórmulas rotas)",
        "• Espacios extra antes/después del texto (usa ESPACIOS() para limpiar)",
        "• Fechas guardadas como texto (no se pueden sumar ni filtrar por fecha)",
    ]
    for i, err in enumerate(errors, 14):
        ws3.cell(row=i, column=1, value=err).font = Font(size=10, color="B71C1C")
        row_height(ws3, i, 16)

    # ── Hoja 4: Formato ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("04 Formato")
    section_title(ws4, 1, 1, "🎨 MÓDULO 4 — Formato y Presentación Profesional", "1F3864")
    note_cell(ws4, 2, 1, "El formato debe comunicar, no decorar. Menos es más.")

    header_row(ws4, 4, ["Técnica de formato", "Cómo aplicarlo", "Cuándo usarlo"], "7B1D32")
    fmt_tips = [
        ("Tabla (Ctrl+T)", "Seleccionar datos → Insertar → Tabla", "Siempre que tengas una lista de datos estructurados"),
        ("Formato condicional", "Inicio → Estilos → Formato condicional", "Resaltar automáticamente valores importantes"),
        ("Formato número personalizado", "Ctrl+1 → Número → Personalizado", 'Ej: #,##0.00 "USD" para mostrar moneda'),
        ("Inmovilizar paneles", "Vista → Inmovilizar → Inmovilizar fila superior", "Cuando tienes muchas filas y necesitas ver encabezados"),
        ("Alto/ancho automático", "Doble clic en borde de columna/fila", "Para ajustar automáticamente al contenido"),
        ("Estilos de celda", "Inicio → Estilos de celda", "Dar formato consistente con un clic"),
    ]
    for i, (t, c, w) in enumerate(fmt_tips, 5):
        data_cell(ws4, i, 1, t, bold=True)
        data_cell(ws4, i, 2, c)
        data_cell(ws4, i, 3, w)
        row_height(ws4, i, 20)

    col_width(ws4, "A", 28); col_width(ws4, "B", 40); col_width(ws4, "C", 45)


def _nivel0_practica(wb, lang):
    ws = wb.active
    ws.title = "PRACTICA Nivel 0"
    section_title(ws, 1, 1, "📝 PRÁCTICA NIVEL 0 — Fundamentos de Excel", "1F3864")
    note_cell(ws, 2, 1, "Completa los ejercicios. Las celdas amarillas son las que debes rellenar.")

    # Ejercicio 1
    section_title(ws, 4, 1, "EJERCICIO 1: Organizar y formatear una tabla de contactos", "1565C0", colspan=5)
    note_cell(ws, 5, 1, "Instrucción: Aplica formato de tabla, ajusta ancho de columnas y agrega colores al encabezado.")
    header_row(ws, 6, ["Nombre","Apellido","Email","Teléfono","Ciudad"], "1976D2")
    contacts = [
        ("Laura","Ramírez","laura@email.com","555-0101","Ciudad de México"),
        ("Pedro","Jiménez","pedro@email.com","555-0102","Bogotá"),
        ("Sofia","Chen","sofia@email.com","555-0103","Lima"),
        ("Andrés","Morales","andres@email.com","555-0104","Buenos Aires"),
        ("Valentina","Soto","valen@email.com","555-0105","Santiago"),
    ]
    for i, row in enumerate(contacts, 7):
        for j, v in enumerate(row, 1):
            data_cell(ws, i, j, v)

    # Ejercicio 2
    section_title(ws, 14, 1, "EJERCICIO 2: Identificar tipos de datos", "2E7D32")
    note_cell(ws, 15, 1, "Instrucción: En la columna B, escribe el tipo de dato (Texto / Número / Fecha / Booleano / Error).")
    header_row(ws, 16, ["Valor en celda", "Tipo de dato (escríbelo tú)", "¿Correcto?"], "43A047")
    examples = ["Excel 365", "365", "3.14", "2024-01-15", "VERDADERO", "#¡DIV/0!", "  espacios  ", "=2+2"]
    for i, ex in enumerate(examples, 17):
        data_cell(ws, i, 1, ex)
        c = data_cell(ws, i, 2, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        data_cell(ws, i, 3, "")

    # Ejercicio 3
    section_title(ws, 27, 1, "EJERCICIO 3: Aplicar atajos de teclado", "E65100")
    note_cell(ws, 28, 1, "Instrucción: Sin usar el mouse, realiza las siguientes acciones en este libro.")
    tasks_kb = [
        ("Ir a la celda A1 usando solo el teclado", "Ctrl + Inicio"),
        ("Seleccionar toda la hoja", "Ctrl + A"),
        ("Poner en negrita el texto de la fila 6", "Ctrl + N"),
        ("Abrir formato de celdas", "Ctrl + 1"),
        ("Ir a la última celda con datos", "Ctrl + Fin"),
    ]
    header_row(ws, 29, ["Tarea", "Atajo correcto"], "EF6C00")
    for i, (task, shortcut) in enumerate(tasks_kb, 30):
        data_cell(ws, i, 1, task)
        data_cell(ws, i, 2, shortcut, bold=True, center_align=True)

    col_width(ws, "A", 32); col_width(ws, "B", 28); col_width(ws, "C", 20)
    col_width(ws, "D", 18); col_width(ws, "E", 20)


# ══════════════════════════════════════════════════════════════════════════════
# NIVEL 1 — Básico
# ══════════════════════════════════════════════════════════════════════════════

def _nivel1_ejemplos(wb, lang):
    F = lambda n: f(lang, n)

    # ── Hoja 1: Tipos de datos y fórmulas básicas ─────────────────────────────
    ws = wb.active
    ws.title = "01 Tipos y Fórmulas"
    section_title(ws, 1, 1, "🔢 MÓDULOS 1-2 — Tipos de Datos y Fórmulas Básicas", "1F3864")

    header_row(ws, 3, ["Descripción", "Fórmula / Ejemplo", "Resultado"], "1565C0")
    col_width(ws, "A", 36); col_width(ws, "B", 38); col_width(ws, "C", 18)

    examples = [
        ("Suma simple", f"={F('SUM')}(10,20,30)", None),
        ("Resta", "=100-45", None),
        ("Multiplicación", "=12*7.5", None),
        ("División", "=250/4", None),
        ("Potencia", "=2^10", None),
        ("Porcentaje de un valor", "=500*15%", None),
        ("Concatenar texto y número", '="Ventas: "&TEXT(1234.5,"#,##0.00")', None),
        ("Fecha de hoy", f"={F('TODAY')}()", None),
        ("Texto en mayúsculas", f'={F("UPPER")}("hola mundo")', None),
        ("Longitud de texto", f'={F("LEN")}("Excel 365")', None),
    ]
    for i, (desc, formula, _) in enumerate(examples, 4):
        formula_example_row(ws, i, desc, formula, formula)
        row_height(ws, i, 18)

    # ── Hoja 2: Referencias ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("02 Referencias")
    section_title(ws2, 1, 1, "🔗 MÓDULO 3 — Referencias: Relativa, Absoluta y Mixta", "1F3864")

    note_cell(ws2, 2, 1, "Tabla de precios con IVA. Observa cómo $C$2 es absoluta (no cambia al copiar la fórmula).")
    ws2.cell(row=2, column=3, value="Tasa IVA:").font = bold_font()
    ws2.cell(row=2, column=4, value=0.16).number_format = "0%"
    ws2.cell(row=2, column=4).font = bold_font(12, "C62828")

    header_row(ws2, 4, ["Producto","Precio base","IVA (ref. absoluta)","Precio final","Fórmula usada"], "1976D2")
    products = ["Laptop","Monitor","Teclado","Mouse","Webcam","Auriculares"]
    prices   = [850, 320, 120, 45, 75, 65]
    for i, (prod, price) in enumerate(zip(products, prices), 5):
        ws2.cell(row=i, column=1, value=prod).border = thin_border()
        ws2.cell(row=i, column=2, value=price).border = thin_border()
        ws2.cell(row=i, column=2).number_format = "#,##0.00"
        ws2.cell(row=i, column=3, value=f"=B{i}*$D$2").border = thin_border()
        ws2.cell(row=i, column=3).number_format = "#,##0.00"
        ws2.cell(row=i, column=4, value=f"=B{i}+C{i}").border = thin_border()
        ws2.cell(row=i, column=4).number_format = "#,##0.00"
        ws2.cell(row=i, column=5, value=f'=B{i}*(1+$D$2)  ← $D$2 no cambia').border = thin_border()
        ws2.cell(row=i, column=5).font = Font(name="Courier New", size=9, color="1A237E")
        row_height(ws2, i, 18)

    col_width(ws2, "A", 16); col_width(ws2, "B", 14); col_width(ws2, "C", 20)
    col_width(ws2, "D", 20); col_width(ws2, "E", 40)

    # ── Hoja 3: Funciones esenciales ──────────────────────────────────────────
    ws3 = wb.create_sheet("03 Funciones Esenciales")
    section_title(ws3, 1, 1, "🧮 MÓDULO 4 — Funciones Esenciales", "1F3864")

    header_row(ws3, 3, ["Función","Sintaxis","¿Qué hace?","Ejemplo"], "1565C0")
    funcs = [
        (F("SUM"),        f"{F('SUM')}(rango)",                    "Suma todos los valores",               f"={F('SUM')}(10,20,30)"),
        (F("AVERAGE"),    f"{F('AVERAGE')}(rango)",                "Calcula el promedio",                  f"={F('AVERAGE')}(10,20,30)"),
        (F("COUNT"),      f"{F('COUNT')}(rango)",                  "Cuenta celdas con números",            f"={F('COUNT')}(A1:A10)"),
        (F("COUNTA"),     f"{F('COUNTA')}(rango)",                 "Cuenta celdas no vacías",              f"={F('COUNTA')}(A1:A10)"),
        (F("MAX"),        f"{F('MAX')}(rango)",                    "Valor máximo",                         f"={F('MAX')}(A1:A10)"),
        (F("MIN"),        f"{F('MIN')}(rango)",                    "Valor mínimo",                         f"={F('MIN')}(A1:A10)"),
        (F("IF"),         f"{F('IF')}(prueba, si_verdadero, si_falso)", "Condición simple",            f'={F("IF")}(A2>100,"Alto","Bajo")'),
        (F("IFERROR"),    f"{F('IFERROR')}(valor, si_error)",      "Maneja errores",                       f'={F("IFERROR")}(A2/B2,0)'),
        (F("ROUND"),      f"{F('ROUND')}(número, decimales)",      "Redondea a N decimales",               f"={F('ROUND')}(3.14159,2)"),
        (F("ABS"),        f"{F('ABS')}(número)",                   "Valor absoluto (quita signo negativo)", f"={F('ABS')}(-15)"),
        (F("TRIM"),       f"{F('TRIM')}(texto)",                   "Elimina espacios extra",               f'={F("TRIM")}("  Excel  ")'),
        (F("TEXT"),       f'{F("TEXT")}(valor, "formato")',        "Convierte número a texto con formato", f'={F("TEXT")}(TODAY(),"DD/MM/YYYY")'),
    ]
    for i, row in enumerate(funcs, 4):
        for j, v in enumerate(row, 1):
            c = data_cell(ws3, i, j, v)
            if j == 2:
                c.font = Font(name="Courier New", size=10, color="1A237E")
            if j == 4:
                c.font = Font(name="Courier New", size=10, color="1B5E20")
        row_height(ws3, i, 20)

    col_width(ws3, "A", 20); col_width(ws3, "B", 36); col_width(ws3, "C", 34); col_width(ws3, "D", 36)

    # ── Hoja 4: Formato Condicional ───────────────────────────────────────────
    ws4 = wb.create_sheet("04 Formato Condicional")
    section_title(ws4, 1, 1, "🎨 MÓDULO 5 — Formato Condicional", "1F3864")
    note_cell(ws4, 2, 1, "Observa cómo el color de la celda cambia automáticamente según el valor. Selecciona el rango C5:C14 para ver las reglas.")

    header_row(ws4, 4, ["Vendedor","Mes","Ventas USD","Semáforo","Nota"], "1565C0")
    vendedores = ["Ana","Carlos","María","Luis","Sofía","Pedro","Valentina","Diego","Isabel","Jorge"]
    import random; random.seed(42)
    ventas = [random.randint(1500, 9500) for _ in range(10)]
    for i, (v, venta) in enumerate(zip(vendedores, ventas), 5):
        data_cell(ws4, i, 1, v)
        data_cell(ws4, i, 2, "Enero 2024")
        ws4.cell(row=i, column=3, value=venta).border = thin_border()
        ws4.cell(row=i, column=3).number_format = "$#,##0"
        semaforo = "🟢 Meta lograda" if venta >= 7000 else ("🟡 En proceso" if venta >= 4000 else "🔴 Bajo meta")
        data_cell(ws4, i, 4, semaforo)
        nota = f'={F("IF")}(C{i}>=7000,"Bono: $"&TEXT(C{i}*0.05,"#,##0"),"Sin bono")'
        ws4.cell(row=i, column=5, value=nota).border = thin_border()
        row_height(ws4, i, 18)

    # Scale de color
    rule = ColorScaleRule(
        start_type="min", start_color="F44336",
        mid_type="percentile", mid_value=50, mid_color="FFEB3B",
        end_type="max", end_color="4CAF50"
    )
    ws4.conditional_formatting.add("C5:C14", rule)

    col_width(ws4, "A", 16); col_width(ws4, "B", 16); col_width(ws4, "C", 14)
    col_width(ws4, "D", 22); col_width(ws4, "E", 32)

    # ── Hoja 5: Gráficos básicos ──────────────────────────────────────────────
    ws5 = wb.create_sheet("05 Graficos")
    section_title(ws5, 1, 1, "📊 MÓDULO 7 — Gráficos Básicos", "1F3864")
    note_cell(ws5, 2, 1, "Selecciona el rango A4:B10 e inserta un gráfico de barras para visualizar las ventas por región.")

    header_row(ws5, 4, ["Región","Ventas Q1","Ventas Q2","Ventas Q3"], "1565C0")
    regiones = [("Norte",12500,14200,13800),("Sur",9800,10500,11200),
                ("Este",15600,16800,15900),("Oeste",8400,9100,9800),
                ("Centro",11200,12400,13100),("Internacional",6700,7800,8500)]
    for i, (r, q1, q2, q3) in enumerate(regiones, 5):
        data_cell(ws5, i, 1, r)
        for j, v in enumerate([q1,q2,q3], 2):
            ws5.cell(row=i, column=j, value=v).border = thin_border()
            ws5.cell(row=i, column=j).number_format = "#,##0"

    chart = BarChart()
    chart.type = "col"
    chart.title = "Ventas por Región"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Región"
    chart.style = 10
    data_ref = Reference(ws5, min_col=2, max_col=4, min_row=4, max_row=10)
    cats_ref = Reference(ws5, min_col=1, min_row=5, max_row=10)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws5.add_chart(chart, "F4")

    col_width(ws5, "A", 18); col_width(ws5, "B", 14); col_width(ws5, "C", 14); col_width(ws5, "D", 14)


def _nivel1_practica(wb, lang):
    F = lambda n: f(lang, n)
    ws = wb.active
    ws.title = "PRACTICA Nivel 1"
    section_title(ws, 1, 1, "📝 PRÁCTICA NIVEL 1 — Básico", "1F3864")
    note_cell(ws, 2, 1, "Completa las celdas amarillas con las fórmulas correctas. No modifiques celdas de datos.")

    # Datos de ventas mensuales
    section_title(ws, 4, 1, "DATOS — Registro de Ventas Mensuales (no modificar)", "424242")
    header_row(ws, 5, ["Mes","Producto","Categoría","Vendedor","Unidades","Precio Unit.","Total Venta"], "616161", "FFFFFF")

    sales_data = [
        ("Enero","Laptop Pro","Electrónicos","Ana García",3,850),
        ("Enero","Mouse Ergonómico","Periféricos","Carlos López",8,45),
        ("Febrero","Monitor 27\"","Electrónicos","María Torres",2,320),
        ("Febrero","Teclado Mecánico","Periféricos","Ana García",5,120),
        ("Marzo","Laptop Pro","Electrónicos","Carlos López",4,850),
        ("Marzo","Webcam HD","Periféricos","María Torres",6,75),
        ("Abril","Monitor 27\"","Electrónicos","Ana García",3,320),
        ("Abril","Mouse Ergonómico","Periféricos","Carlos López",12,45),
        ("Mayo","Silla Gamer","Mobiliario","María Torres",2,280),
        ("Mayo","Laptop Pro","Electrónicos","Ana García",5,850),
    ]
    for i, row in enumerate(sales_data, 6):
        for j, v in enumerate(row, 1):
            data_cell(ws, i, j, v)
        ws.cell(row=i, column=7, value=f"=E{i}*F{i}").border = thin_border()

    # Ejercicio 1
    section_title(ws, 18, 1, "EJERCICIO 1: Calcular totales y estadísticas", "1565C0")
    note_cell(ws, 19, 1, f"Escribe la fórmula correcta en cada celda amarilla. Usa referencias al rango G6:G15.")

    tasks1 = [
        ("Total de todas las ventas", f"={F('SUM')}(G6:G15)"),
        ("Promedio de venta por transacción", f"={F('AVERAGE')}(G6:G15)"),
        ("Número de transacciones", f"={F('COUNT')}(G6:G15)"),
        ("Venta más alta", f"={F('MAX')}(G6:G15)"),
        ("Venta más baja", f"={F('MIN')}(G6:G15)"),
        ("Total redondeado a decena", f"={F('ROUND')}({F('SUM')}(G6:G15),-1)"),
    ]
    header_row(ws, 20, ["Tarea","Tu fórmula (escríbela aquí)","Resultado esperado"], "1976D2")
    for i, (task, formula) in enumerate(tasks1, 21):
        data_cell(ws, i, 1, task)
        c = data_cell(ws, i, 2, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        ws.cell(row=i, column=3, value=formula).font = Font(name="Courier New", size=10, color="388E3C")
        row_height(ws, i, 18)

    # Ejercicio 2
    section_title(ws, 29, 1, "EJERCICIO 2: Fórmulas condicionales", "2E7D32")
    note_cell(ws, 30, 1, "Agrega una columna 'Comisión' en la columna H. La comisión es: 5% si venta > 2000, sino 2%.")
    header_row(ws, 31, ["Fila","Total Venta (ref)","Fórmula SI para comisión","Resultado"], "43A047")
    for i, row_num in enumerate(range(6, 16), 32):
        data_cell(ws, i, 1, row_num)
        data_cell(ws, i, 2, f"=G{row_num}")
        c = data_cell(ws, i, 3, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        expected = f'={F("IF")}(G{row_num}>2000,G{row_num}*5%,G{row_num}*2%)'
        ws.cell(row=i, column=4, value=expected).font = Font(name="Courier New", size=9, color="388E3C")
        row_height(ws, i, 18)

    col_width(ws, "A", 14); col_width(ws, "B", 18); col_width(ws, "C", 18)
    col_width(ws, "D", 14); col_width(ws, "E", 14); col_width(ws, "F", 14); col_width(ws, "G", 16)


# ══════════════════════════════════════════════════════════════════════════════
# NIVEL 2 — Intermedio
# ══════════════════════════════════════════════════════════════════════════════

def _nivel2_ejemplos(wb, lang):
    F = lambda n: f(lang, n)

    ws = wb.active
    ws.title = "01 Condicionales"
    section_title(ws, 1, 1, "🔀 MÓDULOS 2-3 — Funciones Condicionales y Agregación", "1F3864")

    # Tabla de datos para los ejemplos
    note_cell(ws, 2, 1, "Tabla de ventas para practicar SUMIFS, COUNTIFS, AVERAGEIFS y funciones IF avanzadas.")
    header_row(ws, 4, ["Región","Vendedor","Categoría","Mes","Ventas","Meta","¿Meta?","Bono"], "1565C0")

    data = [
        ("Norte","Ana","Electronics","Enero",8500,7000),
        ("Sur","Carlos","Ropa","Enero",4200,5000),
        ("Norte","Ana","Electronics","Febrero",9100,7000),
        ("Este","María","Hogar","Enero",6800,6000),
        ("Sur","Carlos","Electronics","Febrero",7200,7000),
        ("Oeste","Luis","Ropa","Enero",3800,5000),
        ("Norte","María","Hogar","Febrero",5500,6000),
        ("Este","Ana","Electronics","Febrero",11200,7000),
        ("Sur","Luis","Electronics","Enero",6900,7000),
        ("Oeste","Carlos","Hogar","Febrero",7800,6000),
        ("Norte","Luis","Ropa","Enero",5100,5000),
        ("Este","María","Electronics","Enero",8900,7000),
    ]
    for i, row in enumerate(data, 5):
        for j, v in enumerate(row, 1):
            c = data_cell(ws, i, j, v)
        ws.cell(row=i, column=7, value=f'={F("IF")}(E{i}>=F{i},"✅ Sí","❌ No")').border = thin_border()
        ws.cell(row=i, column=8, value=f'={F("IF")}(E{i}>=F{i},E{i}*0.08,0)').border = thin_border()
        ws.cell(row=i, column=8).number_format = "$#,##0"

    # Tabla de resumen con SUMIFS / COUNTIFS
    section_title(ws, 20, 1, f"RESUMEN — Con {F('SUMIFS')} y {F('COUNTIFS')}", "1565C0")
    header_row(ws, 21, ["Región","Total ventas","# Transacciones","Promedio"], "1976D2")

    regiones = ["Norte","Sur","Este","Oeste"]
    for i, reg in enumerate(regiones, 22):
        data_cell(ws, i, 1, reg, bold=True)
        ws.cell(row=i, column=2,
                value=f'={F("SUMIFS")}($E$5:$E$16,$A$5:$A$16,A{i})').border = thin_border()
        ws.cell(row=i, column=2).number_format = "$#,##0"
        ws.cell(row=i, column=3,
                value=f'={F("COUNTIFS")}($A$5:$A$16,A{i})').border = thin_border()
        ws.cell(row=i, column=4,
                value=f'={F("IFERROR")}({F("AVERAGEIFS")}($E$5:$E$16,$A$5:$A$16,A{i}),0)').border = thin_border()
        ws.cell(row=i, column=4).number_format = "$#,##0"

    col_width(ws, "A", 12); col_width(ws, "B", 14); col_width(ws, "C", 14)
    col_width(ws, "D", 12); col_width(ws, "E", 12); col_width(ws, "F", 10)
    col_width(ws, "G", 12); col_width(ws, "H", 14)

    # ── Hoja 2: Funciones de texto ────────────────────────────────────────────
    ws2 = wb.create_sheet("02 Texto")
    section_title(ws2, 1, 1, f"🔤 MÓDULO 4 — Funciones de Texto", "1F3864")
    note_cell(ws2, 2, 1, "Datos crudos importados con problemas de formato. Usa funciones de texto para limpiarlos.")

    header_row(ws2, 4, ["Datos crudos (col A)","Descripción","Fórmula","Resultado"], "4A148C")
    txt_examples = [
        ("  juan pérez  ",  f"Eliminar espacios",          f'={F("TRIM")}(A5)',                 None),
        ("  juan pérez  ",  f"Mayúsculas",                  f'={F("UPPER")}({F("TRIM")}(A6))',   None),
        ("juan pérez",      f"Nombre propio",               f'={F("PROPER")}(A7)',               None),
        ("PRODUCTO-001-MX", f"Extraer primeros 8 chars",    f'={F("LEFT")}(A8,8)',               None),
        ("PRODUCTO-001-MX", f"Extraer código país (últimos 2)", f'={F("RIGHT")}(A8,2)',          None),
        ("PRODUCTO-001-MX", f"Extraer '001' (posición 10, largo 3)", f'={F("MID")}(A10,10,3)',  None),
        ("juan pérez",      f"Largo del texto",             f'={F("LEN")}(A11)',                 None),
        ("hola@mundo.com",  f"Posición del @",              f'={F("FIND")}("@",A12)',            None),
        ("hola@mundo.com",  f"Usuario (antes del @)",       f'={F("LEFT")}(A13,{F("FIND")}("@",A13)-1)', None),
        ("precio: $1,200",  f"Sustituir $ por nada",        f'={F("SUBSTITUTE")}(A14,"$","")',  None),
        ("Ana;Carlos;María",f"Reemplazar ; por coma",       f'={F("SUBSTITUTE")}(A15,";",", ")', None),
    ]
    raw_vals = ["  juan pérez  ","  juan pérez  ","juan pérez","PRODUCTO-001-MX","PRODUCTO-001-MX",
                "PRODUCTO-001-MX","juan pérez","hola@mundo.com","hola@mundo.com","precio: $1,200","Ana;Carlos;María"]

    for i, ((raw, desc, formula, _), raw_v) in enumerate(zip(txt_examples, raw_vals), 5):
        ws2.cell(row=i, column=1, value=raw_v).border = thin_border()
        data_cell(ws2, i, 2, desc)
        ws2.cell(row=i, column=3, value=formula).font = Font(name="Courier New", size=10, color="1A237E")
        ws2.cell(row=i, column=3).border = thin_border()
        ws2.cell(row=i, column=4, value=formula).border = thin_border()
        ws2.cell(row=i, column=4).fill = PatternFill("solid", fgColor="F1F8E9")

    col_width(ws2, "A", 24); col_width(ws2, "B", 36); col_width(ws2, "C", 48); col_width(ws2, "D", 24)

    # ── Hoja 3: Funciones de fecha ────────────────────────────────────────────
    ws3 = wb.create_sheet("03 Fechas")
    section_title(ws3, 1, 1, f"📅 MÓDULO 5 — Funciones de Fecha", "1F3864")
    note_cell(ws3, 2, 1, "Registro de proyectos para practicar funciones de fecha.")

    header_row(ws3, 4, ["Proyecto","Fecha inicio","Fecha fin","Días totales","Días laborables","Mes inicio","Año","¿Terminado?"], "1565C0")
    projects = [
        ("Proyecto Alpha", "2024-01-15", "2024-03-20"),
        ("Proyecto Beta",  "2024-02-01", "2024-04-15"),
        ("Proyecto Gamma", "2024-03-10", "2024-05-30"),
        ("Proyecto Delta", "2024-04-01", "2024-06-01"),
        ("Proyecto Epsilon","2024-05-15", "2024-07-20"),
    ]
    from datetime import date
    today_str = date.today().isoformat()
    for i, (proj, start, end) in enumerate(projects, 5):
        data_cell(ws3, i, 1, proj, bold=True)
        ws3.cell(row=i, column=2, value=start).border = thin_border()
        ws3.cell(row=i, column=2).number_format = "DD/MM/YYYY"
        ws3.cell(row=i, column=3, value=end).border = thin_border()
        ws3.cell(row=i, column=3).number_format = "DD/MM/YYYY"
        ws3.cell(row=i, column=4, value=f"=C{i}-B{i}").border = thin_border()
        ws3.cell(row=i, column=5, value=f"={F('NETWORKDAYS')}(B{i},C{i})").border = thin_border()
        ws3.cell(row=i, column=6, value=f"={F('MONTH')}(B{i})").border = thin_border()
        ws3.cell(row=i, column=7, value=f"={F('YEAR')}(B{i})").border = thin_border()
        ws3.cell(row=i, column=8, value=f'={F("IF")}(C{i}<{F("TODAY")}(),"✅ Sí","⏳ No")').border = thin_border()
        row_height(ws3, i, 18)

    col_width(ws3, "A", 18); col_width(ws3, "B", 14); col_width(ws3, "C", 14)
    col_width(ws3, "D", 14); col_width(ws3, "E", 18); col_width(ws3, "F", 12)
    col_width(ws3, "G", 10); col_width(ws3, "H", 14)

    # ── Hoja 4: Validación de datos ───────────────────────────────────────────
    ws4 = wb.create_sheet("04 Validacion")
    section_title(ws4, 1, 1, "✅ MÓDULO 6 — Validación de Datos", "1F3864")
    note_cell(ws4, 2, 1, "Las columnas con fondo azul claro tienen validación de datos. Intenta escribir un valor inválido.")

    header_row(ws4, 4, ["Empleado","Departamento","Calificación (1-10)","Categoría","Fecha evaluación"], "1565C0")

    # Dropdown validation for departamento
    dv_dept = DataValidation(type="list", formula1='"Ventas,Marketing,Tecnología,RRHH,Finanzas"', allow_blank=True)
    dv_dept.error = "Selecciona un departamento válido de la lista"
    dv_dept.errorTitle = "Departamento inválido"
    ws4.add_data_validation(dv_dept)

    # Number validation for calificacion
    dv_cal = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
    dv_cal.error = "La calificación debe ser un número entero entre 1 y 10"
    dv_cal.errorTitle = "Valor inválido"
    ws4.add_data_validation(dv_cal)

    employees = ["Ana García","Carlos López","María Torres","Luis Morales","Sofía Chen","Pedro Reyes"]
    depts = ["Ventas","Tecnología","Marketing","RRHH","Tecnología","Ventas"]
    scores = [8, 9, 7, 6, 10, 8]
    dates = ["2024-01-15","2024-01-16","2024-01-17","2024-01-18","2024-01-19","2024-01-20"]

    for i, (emp, dept, score, d) in enumerate(zip(employees, depts, scores, dates), 5):
        data_cell(ws4, i, 1, emp)
        c2 = ws4.cell(row=i, column=2, value=dept)
        c2.border = thin_border()
        c2.fill = PatternFill("solid", fgColor="E3F2FD")
        dv_dept.sqref += f"B{i}"
        c3 = ws4.cell(row=i, column=3, value=score)
        c3.border = thin_border()
        c3.fill = PatternFill("solid", fgColor="E3F2FD")
        dv_cal.sqref += f"C{i}"
        ws4.cell(row=i, column=4, value=f'={F("IF")}(C{i}>=9,"Excelente",{F("IF")}(C{i}>=7,"Bueno",{F("IF")}(C{i}>=5,"Regular","Bajo")))').border = thin_border()
        ws4.cell(row=i, column=5, value=d).border = thin_border()
        ws4.cell(row=i, column=5).number_format = "DD/MM/YYYY"

    col_width(ws4, "A", 20); col_width(ws4, "B", 18); col_width(ws4, "C", 20)
    col_width(ws4, "D", 14); col_width(ws4, "E", 18)


def _nivel2_practica(wb, lang):
    F = lambda n: f(lang, n)
    ws = wb.active
    ws.title = "PRACTICA Nivel 2"
    section_title(ws, 1, 1, "📝 PRÁCTICA NIVEL 2 — Intermedio", "1F3864")
    note_cell(ws, 2, 1, "Completa las celdas amarillas. Las fórmulas de referencia están en verde para que puedas verificar.")

    # Datos de recursos humanos
    section_title(ws, 4, 1, "DATOS — Nómina de Empleados (no modificar)", "424242")
    header_row(ws, 5, ["ID","Nombre","Depto","Puesto","Salario","Fecha Ingreso","Antigüedad","País"], "616161")
    employees = [
        ("E001","Ana García","Ventas","Gerente",4500,"2019-03-15","","México"),
        ("E002","Carlos López","TI","Desarrollador",3800,"2021-07-01","","Colombia"),
        ("E003","María Torres","Marketing","Analista",2900,"2020-11-20","","Argentina"),
        ("E004","Luis Morales","Ventas","Ejecutivo",2200,"2022-01-10","","México"),
        ("E005","Sofía Chen","TI","Senior Dev",5200,"2018-05-08","","Colombia"),
        ("E006","Pedro Reyes","RRHH","Coordinador",3100,"2021-09-15","","Perú"),
        ("E007","Valentina Soto","Finanzas","Contador",3600,"2020-04-22","","Chile"),
        ("E008","Diego Moreno","Ventas","Ejecutivo",2100,"2023-02-01","","México"),
    ]
    for i, row in enumerate(employees, 6):
        for j, v in enumerate(row, 1):
            data_cell(ws, i, j, v)
        ws.cell(row=i, column=7, value=f'={F("YEAR")}({F("TODAY")}())-{F("YEAR")}(F{i})').border = thin_border()
        row_height(ws, i, 18)

    # Ejercicio 1: SUMIFS
    section_title(ws, 16, 1, f"EJERCICIO 1: {F('SUMIFS')} — Resumen de nómina por departamento", "1565C0")
    note_cell(ws, 17, 1, f"Usa {F('SUMIFS')} para calcular el total de salarios por departamento.")
    header_row(ws, 18, ["Departamento","Total Salarios (tu fórmula)","Resultado esperado","# Empleados"], "1976D2")
    deptos = [("Ventas", "$C$6:$C$13"), ("TI", "$C$6:$C$13"), ("Marketing", "$C$6:$C$13"), ("Finanzas", "$C$6:$C$13")]
    for i, (dept, _) in enumerate(deptos, 19):
        data_cell(ws, i, 1, dept, bold=True)
        c = data_cell(ws, i, 2, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        expected = f'={F("SUMIFS")}($E$6:$E$13,$C$6:$C$13,A{i})'
        ws.cell(row=i, column=3, value=expected).font = Font(name="Courier New", size=9, color="388E3C")
        ws.cell(row=i, column=4, value=f'={F("COUNTIFS")}($C$6:$C$13,A{i})').border = thin_border()
        row_height(ws, i, 18)

    # Ejercicio 2: Texto
    section_title(ws, 25, 1, "EJERCICIO 2: Manipulación de texto — Crear ID de email corporativo", "2E7D32")
    note_cell(ws, 26, 1, 'Crea el email corporativo: primera letra del nombre + apellido + "@empresa.com". Todo en minúsculas.')
    header_row(ws, 27, ["Nombre completo", "Email corporativo (tu fórmula)", "Resultado esperado"], "43A047")
    full_names = ["Ana García","Carlos López","María Torres","Luis Morales","Sofía Chen"]
    for i, name in enumerate(full_names, 28):
        data_cell(ws, i, 1, name)
        c = data_cell(ws, i, 2, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        # =LOWER(LEFT(A28,1)&MID(A28,FIND(" ",A28)+1,100)&"@empresa.com")
        expected = f'={F("LOWER")}({F("LEFT")}(A{i},1)&{F("MID")}(A{i},{F("FIND")}(" ",A{i})+1,100)&"@empresa.com")'
        ws.cell(row=i, column=3, value=expected).font = Font(name="Courier New", size=9, color="388E3C")
        row_height(ws, i, 18)

    col_width(ws, "A", 14); col_width(ws, "B", 20); col_width(ws, "C", 18)
    col_width(ws, "D", 14); col_width(ws, "E", 12); col_width(ws, "F", 16)
    col_width(ws, "G", 14); col_width(ws, "H", 12)


# ══════════════════════════════════════════════════════════════════════════════
# NIVEL 3 — Avanzado
# ══════════════════════════════════════════════════════════════════════════════

def _nivel3_ejemplos(wb, lang):
    F = lambda n: f(lang, n)

    ws = wb.active
    ws.title = "01 XLOOKUP"
    section_title(ws, 1, 1, f"🔍 MÓDULO 1-2 — {F('XLOOKUP')} e {F('INDEX')}/{F('MATCH')}", "1F3864")
    note_cell(ws, 2, 1, f"{F('XLOOKUP')} es la versión mejorada de BUSCARV/VLOOKUP. Funciona en cualquier dirección.")

    # Catálogo de productos
    section_title(ws, 4, 1, "CATÁLOGO DE PRODUCTOS (tabla de búsqueda)", "424242")
    header_row(ws, 5, ["SKU","Nombre","Categoría","Precio","Stock","Proveedor"], "616161")
    catalog = [
        ("SKU-001","Laptop Pro 15","Electronics",1299,"45","TechSupply"),
        ("SKU-002","Mouse Inalámbrico","Periféricos",39.99,"230","PerifCo"),
        ("SKU-003","Monitor 4K 27\"","Electronics",549,"28","TechSupply"),
        ("SKU-004","Teclado Mecánico","Periféricos",129,"95","PerifCo"),
        ("SKU-005","Webcam 4K","Accesorios",89,"67","AccessPro"),
        ("SKU-006","Hub USB-C 7 puertos","Accesorios",59,"112","AccessPro"),
        ("SKU-007","SSD 1TB","Almacenamiento",119,"78","StoragePlus"),
        ("SKU-008","RAM 32GB DDR5","Componentes",189,"34","TechSupply"),
    ]
    for i, row in enumerate(catalog, 6):
        for j, v in enumerate(row, 1):
            data_cell(ws, i, j, v)

    # Ejemplos XLOOKUP
    section_title(ws, 16, 1, f"EJEMPLOS CON {F('XLOOKUP')}", "1565C0")
    header_row(ws, 17, ["SKU a buscar","¿Qué busco?","Fórmula","Resultado"], "1976D2")

    xlookup_examples = [
        ("SKU-003", "Precio",    f'={F("XLOOKUP")}("SKU-003",$A$6:$A$13,$D$6:$D$13,"No encontrado")'),
        ("SKU-007", "Proveedor", f'={F("XLOOKUP")}("SKU-007",$A$6:$A$13,$F$6:$F$13,"No encontrado")'),
        ("SKU-999", "Precio (no existe — manejo error)", f'={F("XLOOKUP")}("SKU-999",$A$6:$A$13,$D$6:$D$13,"❌ No encontrado")'),
        ("SKU-001", "Nombre + Precio concatenados", f'={F("XLOOKUP")}("SKU-001",$A$6:$A$13,$B$6:$B$13)&" — $"&{F("XLOOKUP")}("SKU-001",$A$6:$A$13,$D$6:$D$13)'),
    ]
    for i, (sku, desc, formula) in enumerate(xlookup_examples, 18):
        data_cell(ws, i, 1, sku, center_align=True)
        data_cell(ws, i, 2, desc)
        ws.cell(row=i, column=3, value=formula).font = Font(name="Courier New", size=9, color="1A237E")
        ws.cell(row=i, column=3).border = thin_border()
        ws.cell(row=i, column=4, value=formula).border = thin_border()
        ws.cell(row=i, column=4).fill = PatternFill("solid", fgColor="F1F8E9")
        row_height(ws, i, 20)

    # INDEX / MATCH
    section_title(ws, 24, 1, f"EJEMPLOS CON {F('INDEX')}/{F('MATCH')}", "E65100")
    note_cell(ws, 25, 1, f"INDEX+MATCH funciona en Excel antiguo. XLOOKUP solo en Excel 365/2019+.")
    header_row(ws, 26, ["Qué busco","Fórmula INDEX/MATCH","Equivalente XLOOKUP"], "EF6C00", "FFFFFF")
    im_examples = [
        ("Precio de SKU-004",
         f'={F("INDEX")}($D$6:$D$13,{F("MATCH")}("SKU-004",$A$6:$A$13,0))',
         f'={F("XLOOKUP")}("SKU-004",$A$6:$A$13,$D$6:$D$13)'),
        ("SKU con mayor stock",
         f'={F("INDEX")}($A$6:$A$13,{F("MATCH")}({F("MAX")}($E$6:$E$13),$E$6:$E$13,0))',
         f'={F("XLOOKUP")}({F("MAX")}($E$6:$E$13),$E$6:$E$13,$A$6:$A$13)'),
        ("Nombre del producto más barato",
         f'={F("INDEX")}($B$6:$B$13,{F("MATCH")}({F("MIN")}($D$6:$D$13),$D$6:$D$13,0))',
         f'={F("XLOOKUP")}({F("MIN")}($D$6:$D$13),$D$6:$D$13,$B$6:$B$13)'),
    ]
    for i, (desc, im, xl) in enumerate(im_examples, 27):
        data_cell(ws, i, 1, desc)
        ws.cell(row=i, column=2, value=im).font = Font(name="Courier New", size=9, color="1A237E")
        ws.cell(row=i, column=2).border = thin_border()
        ws.cell(row=i, column=3, value=xl).font = Font(name="Courier New", size=9, color="1B5E20")
        ws.cell(row=i, column=3).border = thin_border()
        row_height(ws, i, 20)

    col_width(ws, "A", 18); col_width(ws, "B", 30); col_width(ws, "C", 18)
    col_width(ws, "D", 12); col_width(ws, "E", 10); col_width(ws, "F", 16)

    # ── Hoja 2: Arrays dinámicos ──────────────────────────────────────────────
    ws2 = wb.create_sheet("02 Arrays Dinamicos")
    section_title(ws2, 1, 1, f"🌊 MÓDULO 3 — Arrays Dinámicos (Excel 365)", "1F3864")
    note_cell(ws2, 2, 1, f"Las funciones {F('FILTER')}, {F('SORT')} y {F('UNIQUE')} devuelven múltiples resultados automáticamente (se 'derraman').")

    # Datos fuente
    section_title(ws2, 4, 1, "DATOS FUENTE — Transacciones 2024", "424242")
    header_row(ws2, 5, ["Fecha","Cliente","Región","Producto","Monto","Estado"], "616161")
    transactions = [
        ("2024-01-10","TechCorp","Norte","Laptop",2598,"Pagado"),
        ("2024-01-15","StartupXYZ","Sur","Monitor",549,"Pendiente"),
        ("2024-02-01","MegaStore","Norte","Teclado",387,"Pagado"),
        ("2024-02-14","TechCorp","Este","Mouse",199.95,"Pagado"),
        ("2024-03-05","GlobalTech","Oeste","Webcam",356,"Cancelado"),
        ("2024-03-20","StartupXYZ","Norte","Laptop",1299,"Pagado"),
        ("2024-04-01","MegaStore","Sur","SSD",357,"Pagado"),
        ("2024-04-15","TechCorp","Norte","RAM",567,"Pendiente"),
        ("2024-05-10","GlobalTech","Este","Monitor",549,"Pagado"),
        ("2024-05-25","StartupXYZ","Oeste","Hub USB",177,"Pagado"),
    ]
    for i, row in enumerate(transactions, 6):
        for j, v in enumerate(row, 1):
            data_cell(ws2, i, j, v)
            if j == 1:
                ws2.cell(row=i, column=j).number_format = "DD/MM/YYYY"
            if j == 5:
                ws2.cell(row=i, column=j).number_format = "$#,##0.00"

    # Arrays dinámicos
    section_title(ws2, 18, 1, "RESULTADOS CON ARRAYS DINÁMICOS", "1565C0")

    ws2.cell(row=20, column=1, value="Solo transacciones 'Pagado':").font = bold_font(11, "1565C0")
    ws2.cell(row=21, column=1, value=f'={F("FILTER")}($A$6:$F$15,$F$6:$F$15="Pagado","Sin resultados")')
    ws2.cell(row=21, column=1).font = Font(name="Courier New", size=10, color="1A237E")
    ws2.cell(row=21, column=1).border = thin_border()

    ws2.cell(row=28, column=1, value="Clientes únicos (sin repetir):").font = bold_font(11, "2E7D32")
    ws2.cell(row=29, column=1, value=f'={F("UNIQUE")}($B$6:$B$15)')
    ws2.cell(row=29, column=1).font = Font(name="Courier New", size=10, color="1B5E20")
    ws2.cell(row=29, column=1).border = thin_border()

    ws2.cell(row=36, column=1, value="Transacciones ordenadas por monto (mayor a menor):").font = bold_font(11, "E65100")
    ws2.cell(row=37, column=1, value=f'={F("SORT")}($A$6:$F$15,5,-1)')
    ws2.cell(row=37, column=1).font = Font(name="Courier New", size=10, color="BF360C")
    ws2.cell(row=37, column=1).border = thin_border()

    col_width(ws2, "A", 14); col_width(ws2, "B", 16); col_width(ws2, "C", 12)
    col_width(ws2, "D", 16); col_width(ws2, "E", 14); col_width(ws2, "F", 12)


def _nivel3_practica(wb, lang):
    F = lambda n: f(lang, n)
    ws = wb.active
    ws.title = "PRACTICA Nivel 3"
    section_title(ws, 1, 1, "📝 PRÁCTICA NIVEL 3 — Avanzado", "1F3864")
    note_cell(ws, 2, 1, "Usa XLOOKUP, INDEX/MATCH y arrays dinámicos para completar el análisis.")

    # Inventario
    section_title(ws, 4, 1, "DATOS — Inventario de Almacén (no modificar)", "424242")
    header_row(ws, 5, ["Código","Descripción","Categoría","Costo","PVP","Stock","Mín Stock","Estado"], "616161")
    inventory = [
        ("P-001","Disco Duro 2TB","Almacenamiento",45,89,23,10,"OK"),
        ("P-002","Tarjeta Gráfica RTX","GPU",350,699,5,3,"OK"),
        ("P-003","Procesador i9","CPU",280,549,8,5,"OK"),
        ("P-004","Placa Madre Z790","Motherboard",180,329,2,5,"⚠️ BAJO"),
        ("P-005","Fuente 850W","PSU",75,149,15,8,"OK"),
        ("P-006","Gabinete ATX","Carcasa",55,109,3,5,"⚠️ BAJO"),
        ("P-007","Memoria RAM 64GB","RAM",120,219,11,10,"OK"),
        ("P-008","Cooler 360mm","Refrigeración",85,169,6,5,"OK"),
        ("P-009","SSD NVMe 2TB","Almacenamiento",90,179,1,5,"🔴 CRÍTICO"),
        ("P-010","Monitor 144Hz","Display",200,399,18,8,"OK"),
    ]
    for i, row in enumerate(inventory, 6):
        for j, v in enumerate(row, 1):
            data_cell(ws, i, j, v)

    # Ejercicio 1: XLOOKUP lookup table
    section_title(ws, 18, 1, f"EJERCICIO 1: Buscador con {F('XLOOKUP')}", "1565C0")
    note_cell(ws, 19, 1, f"El usuario escribe un código en C20 y debes mostrar todos los datos del producto.")
    ws.cell(row=20, column=1, value="Código a buscar:").font = bold_font()
    c_search = ws.cell(row=20, column=3, value="P-005")
    c_search.fill = PatternFill("solid", fgColor="FFF176")
    c_search.border = thin_border()
    c_search.font = bold_font(12, "1565C0")

    header_row(ws, 22, ["Campo","Valor encontrado","Tu fórmula"], "1976D2")
    fields = ["Descripción","Categoría","Costo","PVP","Stock","Estado"]
    cols   = [2, 3, 4, 5, 6, 8]
    letters = ["B","C","D","E","F","H"]
    for i, (field, col, letter) in enumerate(zip(fields, cols, letters), 23):
        data_cell(ws, i, 1, field, bold=True)
        expected = f'={F("XLOOKUP")}($C$20,$A$6:$A$15,${letter}$6:${letter}$15,"No encontrado")'
        ws.cell(row=i, column=2, value=expected).border = thin_border()
        ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor="F1F8E9")
        c = data_cell(ws, i, 3, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        row_height(ws, i, 18)

    # Ejercicio 2: FILTER
    section_title(ws, 31, 1, f"EJERCICIO 2: {F('FILTER')} — Productos con stock bajo el mínimo", "2E7D32")
    note_cell(ws, 32, 1, f"Usa {F('FILTER')} para mostrar solo los productos donde Stock < Mín Stock.")
    ws.cell(row=33, column=1, value="Tu fórmula aquí:").font = bold_font()
    c_filt = data_cell(ws, 33, 2, "")
    c_filt.fill = PatternFill("solid", fgColor="FFF176")
    expected_filter = f'={F("FILTER")}($A$6:$H$15,$F$6:$F$15<$G$6:$G$15,"Sin productos bajo mínimo")'
    ws.cell(row=34, column=1, value="Respuesta:").font = bold_font(10, "388E3C")
    ws.cell(row=34, column=2, value=expected_filter).font = Font(name="Courier New", size=9, color="388E3C")

    col_width(ws, "A", 14); col_width(ws, "B", 24); col_width(ws, "C", 18)
    col_width(ws, "D", 10); col_width(ws, "E", 10); col_width(ws, "F", 10)
    col_width(ws, "G", 12); col_width(ws, "H", 14)


# ══════════════════════════════════════════════════════════════════════════════
# NIVEL 4 — Excel Pro
# ══════════════════════════════════════════════════════════════════════════════

def _nivel4_ejemplos(wb, lang):
    F = lambda n: f(lang, n)

    ws = wb.active
    ws.title = "01 Macros y VBA"
    section_title(ws, 1, 1, "🤖 MÓDULOS 1-2 — Macros y VBA Básico", "1F3864")
    note_cell(ws, 2, 1, "Las macros automatizan tareas repetitivas. Se guardan en archivos .xlsm (con macros habilitadas).")

    section_title(ws, 4, 1, "GUÍA: Cómo grabar tu primera macro", "4A148C")
    steps = [
        ("Paso 1", "Ir a Vista → Macros → Grabar macro"),
        ("Paso 2", "Asignar un nombre (sin espacios) y atajo de teclado opcional"),
        ("Paso 3", "Guardar en: 'Este libro' para uso personal"),
        ("Paso 4", "Realizar las acciones que quieres automatizar"),
        ("Paso 5", "Ir a Vista → Macros → Detener grabación"),
        ("Paso 6", "Probar con el atajo asignado o Macros → Ejecutar"),
        ("Paso 7", "Guardar el archivo como .xlsm (libro habilitado para macros)"),
    ]
    header_row(ws, 5, ["Paso","Instrucción"], "6A1B9A")
    for i, (step, instr) in enumerate(steps, 6):
        data_cell(ws, i, 1, step, bold=True, center_align=True)
        data_cell(ws, i, 2, instr)
        row_height(ws, i, 20)

    section_title(ws, 15, 1, "EJEMPLOS DE CÓDIGO VBA — Macros básicas", "1F3864")
    vba_examples = [
        ("Mostrar mensaje",
         "Sub MostrarHola()\n    MsgBox \"¡Hola desde VBA!\", vbInformation, \"Bienvenido\"\nEnd Sub"),
        ("Pintar celda activa de amarillo",
         "Sub PintarCelda()\n    ActiveCell.Interior.Color = RGB(255, 255, 0)\n    ActiveCell.Font.Bold = True\nEnd Sub"),
        ("Sumar rango y mostrar resultado",
         "Sub SumarRango()\n    Dim total As Double\n    total = Application.WorksheetFunction.Sum(Range(\"B2:B20\"))\n    MsgBox \"Total: \" & total\nEnd Sub"),
        ("Limpiar celdas vacías en columna A",
         "Sub LimpiarVacios()\n    Dim i As Integer\n    For i = 1 To 100\n        If Cells(i, 1).Value = \"\" Then\n            Cells(i, 1).Interior.Color = RGB(255, 230, 230)\n        End If\n    Next i\nEnd Sub"),
        ("Crear hoja nueva con nombre personalizado",
         "Sub CrearHoja()\n    Dim nombre As String\n    nombre = InputBox(\"Nombre de la nueva hoja:\")\n    If nombre <> \"\" Then\n        Worksheets.Add.Name = nombre\n        MsgBox \"Hoja '\" & nombre & \"' creada!\"\n    End If\nEnd Sub"),
    ]
    header_row(ws, 16, ["Macro","Código VBA (copia esto en el Editor de VBA - Alt+F11)"], "7B1FA2")
    for i, (name, code) in enumerate(vba_examples, 17):
        data_cell(ws, i, 1, name, bold=True)
        c = ws.cell(row=i, column=2, value=code)
        c.font = Font(name="Courier New", size=9, color="1A237E")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = thin_border()
        c.fill = PatternFill("solid", fgColor="F3E5F5")
        ws.row_dimensions[i].height = 70

    col_width(ws, "A", 30); col_width(ws, "B", 70)

    # ── Hoja 2: Dashboard template ────────────────────────────────────────────
    ws2 = wb.create_sheet("02 Dashboard Template")
    section_title(ws2, 1, 1, "📊 MÓDULO — Dashboard Ejecutivo con Fórmulas", "1F3864")
    note_cell(ws2, 2, 1, "Este dashboard se actualiza automáticamente cuando cambias los datos en la hoja '03 Datos'.")

    # KPIs
    section_title(ws2, 4, 1, "KPIs PRINCIPALES — Q1 2024", "1565C0")
    kpis = [
        ("💰 Total Ventas",     f"='{F('SUM')}(Datos!C:C)'",   "$89,450"),
        ("📦 Órdenes",          f"={F('COUNTA')}('03 Datos'!A2:A1000)-1", "247"),
        ("👥 Clientes únicos",  f"=SUMPRODUCT(1/COUNTIF('03 Datos'!B2:B50,'03 Datos'!B2:B50))", "38"),
        ("📈 Ticket promedio",  f"=C5/{F('COUNTA')}('03 Datos'!A2:A248)", "$362"),
    ]
    for i, (kpi, _, val) in enumerate(kpis, 5):
        ws2.cell(row=i, column=1, value=kpi).font = bold_font(12)
        ws2.cell(row=i, column=1).border = thin_border()
        ws2.cell(row=i, column=2, value=val).font = bold_font(18, "1565C0")
        ws2.cell(row=i, column=2).alignment = center()
        ws2.cell(row=i, column=2).border = thin_border()
        ws2.row_dimensions[i].height = 30

    section_title(ws2, 11, 1, "ESTRUCTURA RECOMENDADA PARA UN DASHBOARD", "424242")
    structure = [
        ("1. Hoja de DATOS",     "Datos crudos, nunca modificar formato"),
        ("2. Hoja de CÁLCULOS",  "Tablas resumen, SUMIFS, COUNTIFS"),
        ("3. Hoja de DASHBOARD", "Solo gráficos y KPIs, referenciando hoja de cálculos"),
        ("4. Segmentadores",     "Conectar a tablas dinámicas para filtros interactivos"),
        ("5. Botones de macro",  "Para actualizar datos o navegar entre hojas"),
    ]
    header_row(ws2, 12, ["Capa","Descripción"], "616161")
    for i, (layer, desc) in enumerate(structure, 13):
        data_cell(ws2, i, 1, layer, bold=True)
        data_cell(ws2, i, 2, desc)
        row_height(ws2, i, 20)

    col_width(ws2, "A", 28); col_width(ws2, "B", 55)

    # ── Hoja 3: Datos para dashboard ─────────────────────────────────────────
    ws3 = wb.create_sheet("03 Datos")
    section_title(ws3, 1, 1, "📋 DATOS FUENTE — No modificar estructura", "424242")
    header_row(ws3, 2, ["Orden ID","Cliente","Monto","Producto","Categoría","Región","Fecha","Representante"], "616161")
    import random; random.seed(99)
    clientes = ["TechCorp","StartupXYZ","MegaStore","GlobalTech","InnovateCo","DataSoft","CloudBiz","DigitalPro"]
    productos = ["Laptop","Monitor","Teclado","Mouse","Webcam","SSD","RAM","Hub USB"]
    cats      = ["Electronics","Electronics","Periféricos","Periféricos","Accesorios","Almacenamiento","Componentes","Accesorios"]
    regiones  = ["Norte","Sur","Este","Oeste","Centro"]
    reps      = ["Ana García","Carlos López","María Torres","Luis Morales"]
    from datetime import date, timedelta
    start_date = date(2024, 1, 1)
    for i in range(2, 52):
        row_idx = i + 1
        order_date = start_date + timedelta(days=random.randint(0, 180))
        idx_p = random.randint(0, 7)
        ws3.cell(row=row_idx, column=1, value=f"ORD-{i:04d}").border = thin_border()
        ws3.cell(row=row_idx, column=2, value=random.choice(clientes)).border = thin_border()
        ws3.cell(row=row_idx, column=3, value=round(random.uniform(200, 3000), 2)).border = thin_border()
        ws3.cell(row=row_idx, column=3).number_format = "$#,##0.00"
        ws3.cell(row=row_idx, column=4, value=productos[idx_p]).border = thin_border()
        ws3.cell(row=row_idx, column=5, value=cats[idx_p]).border = thin_border()
        ws3.cell(row=row_idx, column=6, value=random.choice(regiones)).border = thin_border()
        ws3.cell(row=row_idx, column=7, value=order_date.isoformat()).border = thin_border()
        ws3.cell(row=row_idx, column=7).number_format = "DD/MM/YYYY"
        ws3.cell(row=row_idx, column=8, value=random.choice(reps)).border = thin_border()

    for col, width in zip("ABCDEFGH", [14, 16, 14, 16, 16, 10, 14, 18]):
        col_width(ws3, col, width)


def _nivel4_practica(wb, lang):
    F = lambda n: f(lang, n)
    ws = wb.active
    ws.title = "PRACTICA Nivel 4"
    section_title(ws, 1, 1, "📝 PRÁCTICA NIVEL 4 — Excel Pro", "1F3864")
    note_cell(ws, 2, 1, "Desafíos de nivel avanzado. Combina múltiples funciones y construye tu propio dashboard.")

    section_title(ws, 4, 1, "DESAFÍO 1: Función Lambda personalizada (LAMBDA — Excel 365)", "4A148C")
    note_cell(ws, 5, 1, "LAMBDA te permite crear funciones propias. Escribe esto en Administrador de nombres.")
    lambda_examples = [
        ("IVA(monto)",   "=LAMBDA(monto, monto*0.16)",                    "Calcula 16% de IVA"),
        ("PRECIO_FINAL", "=LAMBDA(precio, descuento, precio*(1-descuento))","Precio con descuento"),
        ("DIAS_HABILES",  "=LAMBDA(inicio,fin, DIAS.LAB(inicio,fin))",     "Envuelve DIAS.LAB"),
        ("CATEGORIZAR",  '=LAMBDA(n, IF(n>=9,"A",IF(n>=7,"B","C")))',     "Clasifica nota en A/B/C"),
    ]
    header_row(ws, 6, ["Nombre de la función","Código LAMBDA","Descripción"], "6A1B9A")
    for i, (name, code, desc) in enumerate(lambda_examples, 7):
        data_cell(ws, i, 1, name, bold=True)
        ws.cell(row=i, column=2, value=code).font = Font(name="Courier New", size=10, color="1A237E")
        ws.cell(row=i, column=2).border = thin_border()
        data_cell(ws, i, 3, desc)
        row_height(ws, i, 22)

    section_title(ws, 13, 1, "DESAFÍO 2: Construir un dashboard mínimo viable", "1565C0")
    note_cell(ws, 14, 1, "Crea las siguientes métricas usando los datos de la hoja '03 Datos' del archivo de ejemplos.")
    tasks = [
        ("Top 3 productos por ventas",    f"Combina {F('SORT')} + {F('UNIQUE')} + {F('SUMIFS')}"),
        ("Mapa de calor de ventas por mes y región", "Tabla dinámica + Formato condicional escala de color"),
        ("Alerta de meta no cumplida",    f'{F("IF")}({F("SUM")}(...)< objetivo, "⚠️ ALERTA", "✅ OK")'),
        ("Sparklines en tabla resumen",   "Insertar → Minigráficos → Línea"),
        ("Segmentador de datos",          "Tabla dinámica → Analizar → Insertar segmentador"),
    ]
    header_row(ws, 15, ["Tarea","Técnica sugerida","Estado"], "1976D2")
    for i, (task, tech) in enumerate(tasks, 16):
        data_cell(ws, i, 1, task)
        data_cell(ws, i, 2, tech)
        c = data_cell(ws, i, 3, "⬜ Pendiente")
        c.fill = PatternFill("solid", fgColor="FFF176")
        row_height(ws, i, 22)

    section_title(ws, 23, 1, "DESAFÍO 3: Fórmula anidada compleja — Clasifica clientes", "2E7D32")
    note_cell(ws, 24, 1, "Cliente A: >$5000 total. Cliente B: $2000-$5000. Cliente C: <$2000. Usar SUMIFS + IFS.")
    header_row(ws, 25, ["Cliente","Total compras (dato)","Fórmula de clasificación","Categoría esperada"], "43A047")
    clients_data = [("TechCorp",8500),("StartupXYZ",3200),("GlobalTech",1800),("MegaStore",6100),("DataSoft",2900)]
    for i, (client, total) in enumerate(clients_data, 26):
        data_cell(ws, i, 1, client, bold=True)
        ws.cell(row=i, column=2, value=total).border = thin_border()
        ws.cell(row=i, column=2).number_format = "$#,##0"
        c = data_cell(ws, i, 3, "")
        c.fill = PatternFill("solid", fgColor="FFF176")
        expected = f'={F("IFS")}(B{i}>5000,"⭐ Cliente A",B{i}>=2000,"🟡 Cliente B",B{i}<2000,"Cliente C")'
        ws.cell(row=i, column=4, value=expected).font = Font(name="Courier New", size=9, color="388E3C")
        ws.cell(row=i, column=4).border = thin_border()
        row_height(ws, i, 20)

    col_width(ws, "A", 30); col_width(ws, "B", 36); col_width(ws, "C", 22)


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

_GENERATORS = {
    "nivel0": {"ejemplos": _nivel0_ejemplos, "practica": _nivel0_practica},
    "nivel1": {"ejemplos": _nivel1_ejemplos, "practica": _nivel1_practica},
    "nivel2": {"ejemplos": _nivel2_ejemplos, "practica": _nivel2_practica},
    "nivel3": {"ejemplos": _nivel3_ejemplos, "practica": _nivel3_practica},
    "nivel4": {"ejemplos": _nivel4_ejemplos, "practica": _nivel4_practica},
}

_LEVEL_COLORS = {
    "nivel0": "7B1D32",
    "nivel1": "1565C0",
    "nivel2": "2E7D32",
    "nivel3": "E65100",
    "nivel4": "4A148C",
}

_LEVEL_NAMES = {
    "nivel0": "Nivel 0 — Fundamentos",
    "nivel1": "Nivel 1 — Básico",
    "nivel2": "Nivel 2 — Intermedio",
    "nivel3": "Nivel 3 — Avanzado",
    "nivel4": "Nivel 4 — Excel Pro",
}


def generate_excel(level_key: str, file_type: str, lang: str = "es") -> bytes:
    """
    Generate an Excel file for the given level and type.

    Args:
        level_key: "nivel0" … "nivel4"
        file_type: "ejemplos" or "practica"
        lang: "es" (Spanish formulas) or "en" (English formulas)

    Returns:
        Raw .xlsx bytes ready to send as a file download.
    """
    if lang not in ("es", "en"):
        lang = "es"

    level_key = level_key.lower()
    file_type = file_type.lower()

    generators = _GENERATORS.get(level_key)
    if not generators:
        raise ValueError(f"Unknown level: {level_key}")

    gen_fn = generators.get(file_type)
    if not gen_fn:
        raise ValueError(f"Unknown file type: {file_type}. Use 'ejemplos' or 'practica'.")

    wb = Workbook()
    color = _LEVEL_COLORS.get(level_key, "1F3864")
    level_name = _LEVEL_NAMES.get(level_key, level_key)
    lang_label = "ES" if lang == "es" else "EN"

    # Cover sheet properties
    wb.active.sheet_properties.tabColor = color

    gen_fn(wb, lang)

    # Add a README sheet at the end
    ws_readme = wb.create_sheet("ℹ️ Instrucciones")
    ws_readme.sheet_properties.tabColor = "607D8B"
    section_title(ws_readme, 1, 1, f"📘 {level_name} — Archivo de {file_type.capitalize()} ({lang_label})", "1F3864")
    ws_readme.cell(row=3, column=1, value="Idioma de fórmulas:").font = bold_font()
    ws_readme.cell(row=3, column=2, value="Español" if lang == "es" else "English").font = bold_font(12, color)

    tips = [
        "• Habilita las macros si el archivo lo solicita.",
        "• Las celdas amarillas son las que debes completar en los ejercicios.",
        "• Las celdas verdes muestran la fórmula esperada como referencia.",
        "• Usa Ctrl+` para ver todas las fórmulas del libro.",
        "• Si ves #NAME? intenta cambiar el idioma de fórmulas en tu perfil.",
        "• Guarda como .xlsx para compartir (o .xlsm si contiene macros).",
    ]
    for i, tip in enumerate(tips, 5):
        ws_readme.cell(row=i, column=1, value=tip).font = Font(size=11, color="1F3864")
        row_height(ws_readme, i, 18)

    col_width(ws_readme, "A", 65)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
